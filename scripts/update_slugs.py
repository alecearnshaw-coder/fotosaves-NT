"""
Update Slugs in Species JSON
Automatically populates the Slug field based on old site URL analysis.
Also exports unmatched entries to Excel for manual review.
"""

import json
import csv
import re
from pathlib import Path

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Note: openpyxl not installed. Will create CSV instead of Excel.")
    print("Install with: pip install openpyxl")

# Configuration
COMPARISON_CSV = r"L:\Mis Webs\fotosaves-nt\scripts\slug_comparison_report.csv"
SPECIES_JSON = r"L:\Mis Webs\fotosaves-nt\src\data\taxonomy\species.json"
SPECIES_JSON_BACKUP = r"L:\Mis Webs\fotosaves-nt\src\data\taxonomy\species.json.backup"
UNMATCHED_EXCEL = r"L:\Mis Webs\fotosaves-nt\scripts\unmatched_species.xlsx"
UNMATCHED_CSV = r"L:\Mis Webs\fotosaves-nt\scripts\unmatched_species.csv"

def load_comparison_data():
    """Load the comparison report."""
    data = []
    with open(COMPARISON_CSV, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            data.append(row)
    return data

def load_species_json():
    """Load species JSON."""
    with open(SPECIES_JSON, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_species_json(data):
    """Save species JSON with backup."""
    # Create backup
    with open(SPECIES_JSON, 'r', encoding='utf-8') as f:
        backup_content = f.read()
    with open(SPECIES_JSON_BACKUP, 'w', encoding='utf-8') as f:
        f.write(backup_content)
    print(f"Backup saved to: {SPECIES_JSON_BACKUP}")
    
    # Save updated JSON
    with open(SPECIES_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"Updated JSON saved to: {SPECIES_JSON}")

def export_unmatched_excel(unmatched):
    """Export unmatched entries to Excel."""
    if not EXCEL_AVAILABLE:
        # Fallback to CSV
        with open(UNMATCHED_CSV, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=[
                'old_url', 'old_slug', 'order', 'family', 'notes'
            ])
            writer.writeheader()
            for row in unmatched:
                writer.writerow({
                    'old_url': row['old_url'],
                    'old_slug': row['old_slug'],
                    'order': row['order'],
                    'family': row['family'],
                    'notes': ''
                })
        print(f"Unmatched entries saved to: {UNMATCHED_CSV}")
        return
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Unmatched Species"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Old URL', 'Old Slug', 'Order', 'Family', 'Possible Issue', 'Notes (for manual review)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, row in enumerate(unmatched, 2):
        # Analyze possible issue
        slug = row['old_slug']
        if len(slug) <= 3:
            issue = "Too short - likely incomplete/test"
        elif slug.endswith('X'):
            issue = "Ends with X - likely placeholder"
        elif 'Pato' == slug or 'Pato' in slug and len(slug) < 10:
            issue = "Too generic"
        else:
            issue = "Check species name in JSON"
        
        data = [
            row['old_url'],
            row['old_slug'],
            row['order'],
            row['family'],
            issue,
            ''  # Empty notes column for manual review
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 40
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(UNMATCHED_EXCEL)
    print(f"Unmatched entries saved to: {UNMATCHED_EXCEL}")

def update_slugs():
    """Main function to update slugs."""
    print("Loading data...")
    comparison = load_comparison_data()
    species_data = load_species_json()
    
    # Build update map: species_id -> suggested_slug
    updates = {}
    unmatched = []
    
    for row in comparison:
        if row['status'] == 'NEEDS_UPDATE' and row['species_id']:
            updates[row['species_id']] = row['suggested_slug']
        elif row['status'] == 'NO_MATCH':
            unmatched.append(row)
    
    print(f"  Updates to apply: {len(updates)}")
    print(f"  Unmatched entries: {len(unmatched)}")
    
    # Apply updates
    updated_count = 0
    for species in species_data.get('data', []):
        species_id = species.get('Species_ID')
        if species_id in updates:
            old_slug = species.get('Slug')
            new_slug = updates[species_id]
            species['Slug'] = new_slug
            updated_count += 1
    
    print(f"\nUpdated {updated_count} species records")
    
    # Save updated JSON
    save_species_json(species_data)
    
    # Export unmatched to Excel
    if unmatched:
        print(f"\nExporting {len(unmatched)} unmatched entries...")
        export_unmatched_excel(unmatched)
    
    print("\nDone!")

if __name__ == "__main__":
    update_slugs()


